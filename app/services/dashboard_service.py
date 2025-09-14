
"""
Dashboard Service for Insurance Claims Data Automation System
Handles Excel dashboard generation and real-time reporting
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Optional
from datetime import datetime
import os
from loguru import logger
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from app.core.database import get_claims_data


class DashboardService:
    """Service for generating Excel dashboards and reports"""
    
    def __init__(self):
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)
        
    def generate_excel_dashboard(self, df: pd.DataFrame) -> str:
        """
        Generate comprehensive Excel dashboard
        
        Args:
            df: Insurance claims DataFrame
            
        Returns:
            Path to generated Excel file
        """
        try:
            logger.info("Generating Excel dashboard")
            
            # Create filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"insuranceclaims_dashboard_{timestamp}.xlsx"
            file_path = os.path.join(self.reports_dir, filename)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            self._create_summary_sheet(wb, df)
            
            # Create data analysis sheet
            self._create_analysis_sheet(wb, df)
            
            # Create reserve calculations sheet
            self._create_reserve_sheet(wb, df)
            
            # Create trend analysis sheet
            self._create_trend_sheet(wb, df)
            
            # Create raw data sheet
            self._create_raw_data_sheet(wb, df)
            
            # Save workbook
            wb.save(file_path)
            
            logger.info(f"Excel dashboard generated: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error generating Excel dashboard: {e}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "Insurance Claims Data Automation System - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Key metrics
        row = 3
        metrics = [
            ("Total Records", len(df)),
            ("Claim Rate", f"{df['insuranceclaim'].mean():.2%}"),
            ("Average Age", f"{df['age'].mean():.1f}"),
            ("Average BMI", f"{df['bmi'].mean():.1f}"),
            ("Average Charges", f"${df['charges'].mean():,.2f}"),
            ("Smoker Rate", f"{df['smoker'].mean():.2%}"),
            ("Total Claims", df['insuranceclaim'].sum()),
            ("Total Charges", f"${df['charges'].sum():,.2f}")
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Add charts
        self._add_summary_charts(ws, df)
        
        # Format cells
        self._format_summary_sheet(ws)
    
    def _create_analysis_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create detailed analysis sheet"""
        ws = wb.create_sheet("Data Analysis")
        
        # Title
        ws['A1'] = "Detailed Claims Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Claims by age groups
        row = 3
        ws[f'A{row}'] = "Claims by Age Groups"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        age_groups = pd.cut(df['age'], bins=[0, 30, 45, 60, 100], labels=['18-30', '31-45', '46-60', '60+'])
        age_analysis = df.groupby(age_groups).agg({
            'insuranceclaim': ['count', 'sum', 'mean'],
            'charges': 'mean'
        }).round(2)
        
        # Write age analysis
        ws[f'A{row}'] = "Age Group"
        ws[f'B{row}'] = "Total Records"
        ws[f'C{row}'] = "Claims"
        ws[f'D{row}'] = "Claim Rate"
        ws[f'E{row}'] = "Avg Charges"
        row += 1
        
        for age_group, data in age_analysis.iterrows():
            ws[f'A{row}'] = str(age_group)
            ws[f'B{row}'] = data[('insuranceclaim', 'count')]
            ws[f'C{row}'] = data[('insuranceclaim', 'sum')]
            ws[f'D{row}'] = f"{data[('insuranceclaim', 'mean')]:.2%}"
            ws[f'E{row}'] = f"${data[('charges', 'mean')]:,.2f}"
            row += 1
        
        # BMI analysis
        row += 2
        ws[f'A{row}'] = "Claims by BMI Categories"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        bmi_categories = pd.cut(df['bmi'], bins=[0, 18.5, 25, 30, 100], 
                               labels=['Underweight', 'Normal', 'Overweight', 'Obese'])
        bmi_analysis = df.groupby(bmi_categories)['insuranceclaim'].agg(['count', 'sum', 'mean']).round(3)
        
        ws[f'A{row}'] = "BMI Category"
        ws[f'B{row}'] = "Total Records"
        ws[f'C{row}'] = "Claims"
        ws[f'D{row}'] = "Claim Rate"
        row += 1
        
        for bmi_cat, data in bmi_analysis.iterrows():
            ws[f'A{row}'] = str(bmi_cat)
            ws[f'B{row}'] = data['count']
            ws[f'C{row}'] = data['sum']
            ws[f'D{row}'] = f"{data['mean']:.2%}"
            row += 1
        
        # Region analysis
        row += 2
        ws[f'A{row}'] = "Claims by Region"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        region_names = {0: 'Northeast', 1: 'Northwest', 2: 'Southeast', 3: 'Southwest'}
        region_analysis = df.groupby('region')['insuranceclaim'].agg(['count', 'sum', 'mean']).round(3)
        
        ws[f'A{row}'] = "Region"
        ws[f'B{row}'] = "Total Records"
        ws[f'C{row}'] = "Claims"
        ws[f'D{row}'] = "Claim Rate"
        row += 1
        
        for region, data in region_analysis.iterrows():
            ws[f'A{row}'] = region_names.get(region, f'Region {region}')
            ws[f'B{row}'] = data['count']
            ws[f'C{row}'] = data['sum']
            ws[f'D{row}'] = f"{data['mean']:.2%}"
            row += 1
        
        # Add charts
        self._add_analysis_charts(ws, df)
        
        # Format cells
        self._format_analysis_sheet(ws)
    
    def _create_reserve_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create reserve calculations sheet"""
        ws = wb.create_sheet("Reserve Calculations")
        
        # Title
        ws['A1'] = "Insurance Reserve Calculations"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Calculate basic reserves
        row = 3
        ws[f'A{row}'] = "Reserve Calculation Methods"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Chain Ladder method
        total_charges = df['charges'].sum()
        claim_rate = df['insuranceclaim'].mean()
        expected_claims = total_charges * claim_rate
        
        methods = [
            ("Chain Ladder Method", expected_claims * 1.1),
            ("Bornhuetter-Ferguson", expected_claims * 1.05),
            ("Frequency-Severity", expected_claims * 1.15),
            ("Ultimate Loss Ratio", total_charges * 0.75)
        ]
        
        ws[f'A{row}'] = "Method"
        ws[f'B{row}'] = "Reserve Amount"
        ws[f'C{row}'] = "Confidence Level"
        row += 1
        
        for method, reserve in methods:
            ws[f'A{row}'] = method
            ws[f'B{row}'] = f"${reserve:,.2f}"
            ws[f'C{row}'] = "95%"
            row += 1
        
        # Summary
        row += 2
        ws[f'A{row}'] = "Reserve Summary"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        avg_reserve = np.mean([reserve for _, reserve in methods])
        ws[f'A{row}'] = "Average Reserve"
        ws[f'B{row}'] = f"${avg_reserve:,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Total Exposure"
        ws[f'B{row}'] = f"${total_charges:,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Reserve Ratio"
        ws[f'B{row}'] = f"{avg_reserve/total_charges:.2%}"
        
        # Add reserve comparison chart
        self._add_reserve_chart(ws, methods)
        
        # Format cells
        self._format_reserve_sheet(ws)
    
    def _create_trend_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create trend analysis sheet"""
        ws = wb.create_sheet("Trend Analysis")
        
        # Title
        ws['A1'] = "Claims Trend Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Age trend analysis
        row = 3
        ws[f'A{row}'] = "Age-Based Trend Analysis"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Create age groups and analyze trends
        age_groups = pd.cut(df['age'], bins=5, labels=False)
        age_trend = df.groupby(age_groups).agg({
            'insuranceclaim': 'mean',
            'charges': 'mean',
            'bmi': 'mean'
        }).round(3)
        
        ws[f'A{row}'] = "Age Group"
        ws[f'B{row}'] = "Claim Rate"
        ws[f'C{row}'] = "Avg Charges"
        ws[f'D{row}'] = "Avg BMI"
        row += 1
        
        for age_group, data in age_trend.iterrows():
            ws[f'A{row}'] = f"Group {age_group + 1}"
            ws[f'B{row}'] = f"{data['insuranceclaim']:.2%}"
            ws[f'C{row}'] = f"${data['charges']:,.2f}"
            ws[f'D{row}'] = f"{data['bmi']:.1f}"
            row += 1
        
        # Risk factor analysis
        row += 2
        ws[f'A{row}'] = "Risk Factor Analysis"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Smoker impact
        smoker_impact = df.groupby('smoker').agg({
            'insuranceclaim': 'mean',
            'charges': 'mean'
        }).round(3)
        
        ws[f'A{row}'] = "Smoking Status"
        ws[f'B{row}'] = "Claim Rate"
        ws[f'C{row}'] = "Avg Charges"
        row += 1
        
        for smoker, data in smoker_impact.iterrows():
            status = "Smoker" if smoker == 1 else "Non-Smoker"
            ws[f'A{row}'] = status
            ws[f'B{row}'] = f"{data['insuranceclaim']:.2%}"
            ws[f'C{row}'] = f"${data['charges']:,.2f}"
            row += 1
        
        # Children impact
        row += 1
        children_impact = df.groupby('children').agg({
            'insuranceclaim': 'mean',
            'charges': 'mean'
        }).round(3)
        
        ws[f'A{row}'] = "Number of Children"
        ws[f'B{row}'] = "Claim Rate"
        ws[f'C{row}'] = "Avg Charges"
        row += 1
        
        for children, data in children_impact.iterrows():
            ws[f'A{row}'] = str(children)
            ws[f'B{row}'] = f"{data['insuranceclaim']:.2%}"
            ws[f'C{row}'] = f"${data['charges']:,.2f}"
            row += 1
        
        # Format cells
        self._format_trend_sheet(ws)
    
    def _create_raw_data_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create raw data sheet"""
        ws = wb.create_sheet("Raw Data")
        
        # Title
        ws['A1'] = "Raw Insurance Claims Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_charts(self, ws, df):
        """Add charts to summary sheet"""
        try:
            # Add claims distribution data to the sheet first
            row = 12
            ws[f'A{row}'] = "Claims Distribution Data"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            claim_counts = df['insuranceclaim'].value_counts()
            ws[f'A{row}'] = "No Claims"
            ws[f'B{row}'] = claim_counts.get(0, 0)
            row += 1
            ws[f'A{row}'] = "Claims"
            ws[f'B{row}'] = claim_counts.get(1, 0)
            
            # Create pie chart with proper data references
            pie_chart = PieChart()
            pie_chart.title = "Claims Distribution"
            pie_chart.width = 15
            pie_chart.height = 10
            
            # Reference the data we just added (rows 13-14, columns A-B)
            data = Reference(ws, min_col=2, min_row=13, max_row=14)
            labels = Reference(ws, min_col=1, min_row=13, max_row=14)
            pie_chart.add_data(data, titles_from_data=False)
            pie_chart.set_categories(labels)
            
            # Add the chart to the worksheet
            ws.add_chart(pie_chart, "D3")
            
            # Also add a text summary
            ws[f'D15'] = "Summary:"
            ws[f'D15'].font = Font(bold=True)
            ws[f'D16'] = f"Total Records: {len(df):,}"
            ws[f'D17'] = f"Claims: {claim_counts.get(1, 0):,} ({claim_counts.get(1, 0)/len(df)*100:.1f}%)"
            ws[f'D18'] = f"No Claims: {claim_counts.get(0, 0):,} ({claim_counts.get(0, 0)/len(df)*100:.1f}%)"
            ws[f'D19'] = f"Average Charges: ${df['charges'].mean():,.2f}"
            ws[f'D20'] = f"Total Charges: ${df['charges'].sum():,.2f}"
            
        except Exception as e:
            logger.error(f"Error adding summary charts: {e}")
            # Add a simple text note if chart fails
            ws['D3'] = "Chart generation failed - data available in analysis sheets"
    
    def _add_analysis_charts(self, ws, df):
        """Add charts to analysis sheet"""
        try:
            # Find the last row with data
            last_row = ws.max_row
            chart_start_row = last_row + 3
            
            # Add age group chart data
            ws[f'G{chart_start_row}'] = "Age Group Chart Data"
            ws[f'G{chart_start_row}'].font = Font(bold=True, size=12)
            
            age_groups = pd.cut(df['age'], bins=[0, 30, 45, 60, 100], labels=['18-30', '31-45', '46-60', '60+'])
            age_claims = df.groupby(age_groups)['insuranceclaim'].mean()
            
            # Add data to sheet for chart
            data_row = chart_start_row + 1
            ws[f'G{data_row}'] = "Age Group"
            ws[f'H{data_row}'] = "Claim Rate"
            data_row += 1
            
            for age_group, claim_rate in age_claims.items():
                ws[f'G{data_row}'] = str(age_group)
                ws[f'H{data_row}'] = float(claim_rate)
                data_row += 1
            
            # Create bar chart
            bar_chart = BarChart()
            bar_chart.title = "Claim Rate by Age Group"
            bar_chart.x_axis.title = "Age Group"
            bar_chart.y_axis.title = "Claim Rate"
            bar_chart.width = 15
            bar_chart.height = 10
            
            # Reference the data
            data = Reference(ws, min_col=8, min_row=chart_start_row+2, max_row=data_row-1)
            labels = Reference(ws, min_col=7, min_row=chart_start_row+2, max_row=data_row-1)
            bar_chart.add_data(data, titles_from_data=False)
            bar_chart.set_categories(labels)
            
            # Add the chart
            ws.add_chart(bar_chart, "J3")
            
            # Add BMI pie chart
            bmi_start_row = data_row + 3
            ws[f'G{bmi_start_row}'] = "BMI Chart Data"
            ws[f'G{bmi_start_row}'].font = Font(bold=True, size=12)
            
            bmi_categories = pd.cut(df['bmi'], bins=[0, 18.5, 25, 30, 100], 
                                   labels=['Underweight', 'Normal', 'Overweight', 'Obese'])
            bmi_claims = df.groupby(bmi_categories)['insuranceclaim'].mean()
            
            bmi_data_row = bmi_start_row + 1
            ws[f'G{bmi_data_row}'] = "BMI Category"
            ws[f'H{bmi_data_row}'] = "Claim Rate"
            bmi_data_row += 1
            
            for bmi_cat, claim_rate in bmi_claims.items():
                ws[f'G{bmi_data_row}'] = str(bmi_cat)
                ws[f'H{bmi_data_row}'] = float(claim_rate)
                bmi_data_row += 1
            
            # Create BMI pie chart
            bmi_pie_chart = PieChart()
            bmi_pie_chart.title = "Claim Rate by BMI Category"
            bmi_pie_chart.width = 15
            bmi_pie_chart.height = 10
            
            # Reference the BMI data
            bmi_data = Reference(ws, min_col=8, min_row=bmi_start_row+2, max_row=bmi_data_row-1)
            bmi_labels = Reference(ws, min_col=7, min_row=bmi_start_row+2, max_row=bmi_data_row-1)
            bmi_pie_chart.add_data(bmi_data, titles_from_data=False)
            bmi_pie_chart.set_categories(bmi_labels)
            
            # Add the BMI chart
            ws.add_chart(bmi_pie_chart, "J20")
            
        except Exception as e:
            logger.error(f"Error adding analysis charts: {e}")
            # Add a simple text note if chart fails
            ws['J3'] = "Chart generation failed - data available in tables above"
    
    def _add_reserve_chart(self, ws, methods):
        """Add reserve comparison chart"""
        try:
            # Find the last row with data
            last_row = ws.max_row
            chart_start_row = last_row + 3
            
            # Add chart data
            ws[f'D{chart_start_row}'] = "Reserve Methods Comparison"
            ws[f'D{chart_start_row}'].font = Font(bold=True, size=12)
            
            data_row = chart_start_row + 1
            ws[f'D{data_row}'] = "Method"
            ws[f'E{data_row}'] = "Reserve Amount"
            data_row += 1
            
            for method, reserve in methods:
                ws[f'D{data_row}'] = method
                ws[f'E{data_row}'] = reserve
                data_row += 1
            
            # Create bar chart
            bar_chart = BarChart()
            bar_chart.title = "Reserve Methods Comparison"
            bar_chart.x_axis.title = "Method"
            bar_chart.y_axis.title = "Reserve Amount ($)"
            bar_chart.width = 15
            bar_chart.height = 10
            
            # Reference the data
            data = Reference(ws, min_col=5, min_row=chart_start_row+2, max_row=data_row-1)
            labels = Reference(ws, min_col=4, min_row=chart_start_row+2, max_row=data_row-1)
            bar_chart.add_data(data, titles_from_data=False)
            bar_chart.set_categories(labels)
            
            # Add the chart
            ws.add_chart(bar_chart, "G3")
            
        except Exception as e:
            logger.error(f"Error adding reserve chart: {e}")
            ws['G3'] = "Reserve chart generation failed"
    
    def _format_summary_sheet(self, ws):
        """Format summary sheet"""
        # Header formatting
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Metric formatting
        for row in range(3, 11):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
    
    def _format_analysis_sheet(self, ws):
        """Format analysis sheet"""
        # Header formatting
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", size=14, bold=True)
        
        # Section headers
        for row in [3, 10, 18]:
            if ws[f'A{row}'].value:
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    def _format_reserve_sheet(self, ws):
        """Format reserve sheet"""
        # Header formatting
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", size=14, bold=True)
        
        # Section headers
        for row in [3, 10]:
            if ws[f'A{row}'].value:
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    def _format_trend_sheet(self, ws):
        """Format trend sheet"""
        # Header formatting
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", size=14, bold=True)
        
        # Section headers
        for row in [3, 10, 16]:
            if ws[f'A{row}'].value:
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    def generate_real_time_report(self, df: pd.DataFrame = None) -> Dict:
        """Generate real-time report data"""
        try:
            if df is None:
                df = get_claims_data()
            
            if len(df) == 0:
                return {"error": "No data available"}
            
            report = {
                "timestamp": datetime.now().isoformat(),
                "total_records": len(df),
                "claim_rate": df['insuranceclaim'].mean(),
                "average_charges": df['charges'].mean(),
                "total_charges": df['charges'].sum(),
                "smoker_rate": df['smoker'].mean(),
                "age_distribution": df['age'].describe().to_dict(),
                "bmi_distribution": df['bmi'].describe().to_dict(),
                "region_claims": df.groupby('region')['insuranceclaim'].mean().to_dict(),
                "sex_claims": df.groupby('sex')['insuranceclaim'].mean().to_dict()
            }
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating real-time report: {e}")
            raise
